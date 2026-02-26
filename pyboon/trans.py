"""
table 相关转换
"""


import os
import json
from openpyxl import Workbook, load_workbook
from pyboon import file as f
from rich.console import Console


def csv2xl(csvfile):
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Failed to get active worksheet")

    txt = f.read(csvfile)

    for line in txt.split("\n"):
        if line.strip():  # 跳过空行
            print(line)
            ws.append(line.split(","))

    wb.save(csvfile.replace(".txt", ".xlsx"))


def xl2csv(xlfile):
    wb = load_workbook(xlfile)
    ws = wb[wb.sheetnames[0]]

    lines = []
    for row in ws.iter_rows():
        values = [str(cell.value) if cell.value is not None else "" for cell in row]
        if values[0]:
            lines.append(",".join(values))

    f.write(os.path.join(os.getcwd(), xlfile.replace(
        ".xlsx", ".txt")), "\n".join(lines))


def to_json(d):
    return json.dumps(d, ensure_ascii=False)


def to_json_array(l):
    return [to_json(x) for x in l]


def to_md_line(d, head=[]):
    if not head:
        head = [k for k,_ in d.items()]
    return "| "+" | ".join([str(d[k]) for k in head])+" |"

def get_show_list(d, head=[], show={}):
    if not head:
        head = [k for k, _ in d.items()]
    if not show:
        show_list = head
    else:
        show_list = [show[k] for k in head ]
    return show_list

def to_md(l, head=[], show={}):
    if not l or not l[0]:
        return ""
    show_list = get_show_list(l[0],head,show)
    showline = "| "+" | ".join(show_list)+" |"
    splitline = "| "+" |".join([":---:" for k, _ in l[0].items()])+" |"
    lines = [showline, splitline]
    for row in l:
        line = to_md_line(row, head)
        lines.append(line)
    return "\n".join(lines)


def to_html(l,head=[], show={}):
    md = to_md(l,head,show)
    import markdown
    html = markdown.markdown(md, extensions=['tables'])
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html, 'html.parser')
    return soup.prettify()


def to_csv_line(d, head=[]):
    if not head:
        head = [k for k,_ in d.items()]
    return ",".join([str(d[k]) for k in head])


def to_csv(l, head=[], show=[]):
    if not l or not l[0]:
        return ""
    show_list = get_show_list(l[0],head,show)
    showline = ",".join(show_list)
    lines = [showline]
    for row in l:
        line = to_csv_line(row)
        lines.append(line)
    return "\n".join(lines)


def to_xlsx(l, head=[], show=[]):
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Failed to get active worksheet")

    for line in to_csv(l,head,show).split("\n"):
        if line.strip():  # 跳过空行
            print(line)
            ws.append(line.split(","))

    wb.save("a.xlsx")


if __name__ == '__main__':
    console = Console()
    console.print(to_csv([{"a": 1, "b": "c"}, {"a": 2, "b": "c"}],head=['b','a']))
    console.print(to_xlsx([{"a": 1, "b": "c"}, {"a": 2, "b": "c"}],head=['b','a']))
    console.print(to_md([{"a": 1, "b": "c"}, {"a": 2, "b": "c"}],head=['b','a']))
    console.print(to_html([{"a": 1, "b": "c"}, {"a": 2, "b": "c"}],head=['b','a'],show={"a":"第一","b":"第二",}))
